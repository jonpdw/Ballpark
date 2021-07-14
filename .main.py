import openpyxl as xl
from openpyxl.styles import PatternFill, colors, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.borders import Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles import numbers
import re
import numpy as np
import datetime as dt
import pandas as pd
from collections import namedtuple
from string import ascii_uppercase
from dateutil.parser import parse
import holidays
import glob
import sys
import os

print("Started")

excel_files0 = glob.glob("*.xlsx")
excel_files1 = [x for x in excel_files0 if "~" not in x]
excel_files2 = [x for x in excel_files1 if "Setup" not in x]
excel_files3 = [x for x in excel_files2 if "Ballpark" not in x]
excel_files4 = [x for x in excel_files3 if "Prebooked work" not in x]
excel_files5 = [(parse(x[:-5]), x) for x in excel_files4]
load_wb = sorted(excel_files5, key=lambda x: x[0])[-1][1]

wb = xl.load_workbook(load_wb, data_only=True)
setup = xl.load_workbook("Setup.xlsx", read_only=True)
waiting_for_job = xl.load_workbook("Prebooked work.xlsx", read_only=True)

# only use my manual date if we are on my computer using pycharm
if len(sys.argv) == 2 and sys.argv[1] == "use_manual_date":
    today = dt.date(2020, 11, 20)
else:
    today = dt.date.today()

current_year = dt.datetime.now().year
nz_holidays = list(holidays.CountryHoliday("NZL", years=[current_year, current_year + 1], prov="AUK"))


def get_unit_equiv_from_setup(ws):
    """Reads the Units & Job Type Sheet into a dictionary"""
    dict = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        contains_string, type, one_unit_if_front = row
        dict[contains_string.value] = {"type": type.value, "one_unit_if_front": one_unit_if_front.value}
    return dict


def get_location_department_from_setup(ws):
    """Reads the Location Departmente Sheet into a dictionary"""
    ignore_list = []
    value_dict = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        person, status = row
        if status.value == "Ignore":
            ignore_list.append(person.value)
        else:
            value_dict[person.value] = status.value
    return (ignore_list, value_dict)


def get_not_standard_format_ignore_list_from_setup(ws):
    """Reads the 2 Not Standard Format list from setup"""
    ignore_list = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value is not None:
            ignore_list.append(row[0].value)
    return ignore_list


def get_daily_total_capasity(ws):
    daily_total = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        vals = [x.value for x in row if x.value is not None]
        if len(vals) == 0: continue
        daily_total[vals[0].date()] = sum(vals[1:])
    return daily_total


def max_row_calc(ws):
    row_num = 1
    while (True):
        if ws.cell(row=row_num, column=1).value is None:
            row_num -= 1
            break
        row_num += 1
    return row_num


class NoPatternsMatched(Exception): pass


class NotStandardDescriptionFormat(Exception): pass


class DescriptionNull(Exception): pass


def description_parse(desc_row, last_dep):
    if last_dep == "ON HOLD": return "On Hold"
    if desc_row is None: raise DescriptionNull
    unit_override = re.findall(r"\[(\d+)\]", desc_row)
    if len(unit_override) != 0:
        return int(unit_override[0])

    desc_row = desc_row.replace("REMAKE", "")
    # Sometimes there are multiple jobs in one line that could have different unit amounts e.g. "1 x EMAX 1 x FGC- GOLD (26) (24)"
    results = re.findall(r"\d+\s?[xX]\s(?:3\/4)*[A-Za-z\s/.\-/&]*(?:\([\d,]*\)\s?)*", desc_row)
    if len(results) > 0:
        return sum(map(description_pattern_match, results))
    else:
        raise NotStandardDescriptionFormat


def unit_equiv_from_mouth_position(str1):
    if (str1[0] in ["1", "2"]) and (str1[1] in ["1", "2", "3"]):
        return 1
    else:
        return 0.5


def description_pattern_match(desc_row):
    for job_type_bit_to_contain in unit_equiv_dict:
        if job_type_bit_to_contain in desc_row:
            # if it is a job type e.g e-max thats total depends on whether it is a front tooth e.g. (21) then we need some special logic to do that
            if unit_equiv_dict[job_type_bit_to_contain]["one_unit_if_front"] is True:
                if unit_equiv_dict[job_type_bit_to_contain]["type"] == 0: return 0
                mouth_pos = re.findall("\(([\d,]*)\)", desc_row)
                if len(mouth_pos) != 0:
                    mouth_pos_list = ",".join(mouth_pos).split(",")
                    return sum(map(unit_equiv_from_mouth_position, mouth_pos_list))
            return int(re.findall(r"\d+", desc_row)[0]) * unit_equiv_dict[job_type_bit_to_contain]["type"]
    raise NoPatternsMatched


def get_sheet_for_day(name, wb):
    """Creates a new sheet with column headers or returns one that has been created before"""
    if name not in wb:
        sheet = wb.create_sheet(f"{name}")
        return sheet
    else:
        return wb[name]


def generate_column_dictionarys():
    Column = namedtuple("Columns", ["name", "index1", "letter", "index0"])

    cols_names = ["CaseID", "Dentist", "Patient", "Description", "Total", "Last Department", "Ahead/Behind",
                  "INSERT DATE"]
    cols_names_count = len(cols_names) + 1
    cols_index = list(range(1, cols_names_count))
    cols_letter = list(ascii_uppercase)[0:cols_names_count]

    cols_names_d = {}
    cols_index_d = {}
    cols_letter_d = {}
    for i in range(len(cols_names)):
        cols_names_d[cols_names[i]] = Column(cols_names[i], cols_index[i], cols_letter[i], i)
        cols_index_d[cols_index[i]] = Column(cols_names[i], cols_index[i], cols_letter[i], i)
        cols_letter_d[cols_letter[i]] = Column(cols_names[i], cols_index[i], cols_letter[i], i)
    return cols_names_d, cols_index_d, cols_letter_d, cols_names_count


colNam, colIdx, colLet, colNum = generate_column_dictionarys()
unit_equiv_dict = get_unit_equiv_from_setup(setup["1 Units & Job Type"])
ignore_list, location_dep_dict = get_location_department_from_setup(setup["Location Department"])
not_standard_format_ignore = get_not_standard_format_ignore_list_from_setup(setup["2 Not Standard Format"])
daily_total = get_daily_total_capasity(setup["Dates"])
ws = wb['Events']

# delete first row that doesn't contain any info
ws.delete_rows(1)
ws.delete_cols(2)
wb.remove(wb['Sheet2'])
wb.remove(wb['Sheet3'])

max_row = max_row_calc(ws)

"""Normal rows from sheet"""
for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=7):
    caseID, dentist, patient, desc, last_dep, dateout, insertdate = row
    if last_dep.value == "ON HOLD": continue;
    problem = ""
    ahead_behind = None
    total = 0
    porcelain_date = np.busday_offset(dateout.value.date(), -1, roll="backward", holidays=nz_holidays)
    sheetName = str(np.busday_count(today, porcelain_date, holidays=nz_holidays))
    sheet = get_sheet_for_day(sheetName, wb)
    cur_row = sheet.max_row + 1
    sheet.cell(1, 8).value = dateout.value
    try:
        total = description_parse(desc.value, last_dep.value)
        if total == 0:
            continue
            # total = "Ignore List"
    except NoPatternsMatched:
        # problem = 1
        total = "No Patterns Matched"
    except NotStandardDescriptionFormat:
        # problem = 2
        total = "Not Standard Description Type"
        for x in not_standard_format_ignore:
            if x in desc.value:
                continue
    except DescriptionNull:
        total = "Description Null"

    if last_dep.value is None:
        total = "No last department"
    elif (last_dep.value not in location_dep_dict) and (last_dep.value not in ignore_list):
        total = "Last department not in setup file"
    elif last_dep.value not in ignore_list:
        modifier = int(sheetName)
        if modifier > 2: modifier = 2
        ahead_behind = location_dep_dict[last_dep.value] + modifier

    if last_dep.value == "WAITING FOR SHADE" and sheetName in ["0", "1", "2"]:
        total = "Waiting for shade"

    sheet.cell(cur_row, colNam['CaseID'].index1).value = caseID.value
    sheet.cell(cur_row, colNam['Dentist'].index1).value = dentist.value
    sheet.cell(cur_row, colNam['Patient'].index1).value = patient.value
    sheet.cell(cur_row, colNam['Description'].index1).value = desc.value
    sheet.cell(cur_row, colNam['Total'].index1).value = total
    sheet.cell(cur_row, colNam['Last Department'].index1).value = last_dep.value
    sheet.cell(cur_row, colNam['Ahead/Behind'].index1).value = ahead_behind
    sheet.cell(cur_row, colNam['INSERT DATE'].index1).value = insertdate.value
    # sheet.cell(cur_row, colNam['Problem'].index1).value = problem


"""Add Prebooking"""
waiting_ws = waiting_for_job['Sheet1']

for row in waiting_ws.iter_rows(min_row=2, max_row=waiting_ws.max_row, min_col=1, max_col=5):
    dentist, patient, desc, total, porcelain_date = row
    if (dentist.value is None) and (patient.value is None) and (desc.value is None) and (total.value is None) and (
            porcelain_date.value is None): continue
    porcelain_date = porcelain_date.value.date()
    if porcelain_date < today: continue
    problem = ""
    ahead_behind = None
    sheetName = str(np.busday_count(today, porcelain_date, holidays=nz_holidays))
    if sheetName not in wb: continue
    sheet = get_sheet_for_day(sheetName, wb)
    cur_row = sheet.max_row + 1

    if total.value is None:
        total = 0
        try:
            total = description_parse(desc.value, "")
            if total == 0: total = "Ignore List"
        except NoPatternsMatched:
            total = "No Patterns Matched"
        except NotStandardDescriptionFormat:
            total = "Not Standard Description Type"
            for x in not_standard_format_ignore:
                if x in desc.value:
                    total = "Ignore List"
        except DescriptionNull:
            total = "Description Null"
    else:
        total = total.value

    # sheet.cell(cur_row, colNam['CaseID'].index1).value = caseID.value
    sheet.cell(cur_row, colNam['Dentist'].index1).value = dentist.value
    sheet.cell(cur_row, colNam['Patient'].index1).value = patient.value
    sheet.cell(cur_row, colNam['Description'].index1).value = desc.value
    sheet.cell(cur_row, colNam['Total'].index1).value = total
    sheet.cell(cur_row, colNam['Last Department'].index1).value = "Prebooked"
    # sheet.cell(cur_row, colNam['Ahead/Behind'].index1).value = ahead_behind
    # sheet.cell(cur_row, colNam['INSERT DATE'].index1).value = insertdate.value
    # sheet.cell(cur_row, colNam['Problem'].index1).value = problem

    dateout = np.busday_offset(porcelain_date, 1, roll="forward", holidays=nz_holidays)
    sheet.cell(1, 8).value = pd.to_datetime(dateout)

# """Sort columns"""
# for sheet_name in wb.sheetnames:
#     if sheet_name == "Events": continue
#     sheet = wb[sheet_name]
#     sorted_sheet = wb.create_sheet(sheet_name + " sorting")
#     rows = list(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=colNum))
#     rows_sorted = sorted(rows, key=lambda row: row[colNam['Ahead/Behind'].index0].value if type(
#         row[colNam['Ahead/Behind'].index0].value) is int else 3)
#     for row_idx, row in enumerate(rows_sorted, 3):
#         for col_idx, col in enumerate(row, 1):
#             sorted_sheet.cell(row_idx, col_idx).value = col.value
#     sorted_sheet.cell(1, 8).value = sheet.cell(1, 8).value
#     wb.remove(sheet)
#     sorted_sheet.title = sheet_name


"""Add daily total """
for sheet_name in wb.sheetnames:
    if sheet_name == "Events": continue
    sheet = wb[sheet_name]
    x = colNam['Total'].letter
    cur_date = np.busday_offset(today, int(sheet_name), roll="forward", holidays=nz_holidays)
    cur_date1 = pd.to_datetime(cur_date).date()
    sheet.cell(1, 4).value = daily_total.get(cur_date1, "n/a")
    sheet.cell(1, 4).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(1, 4).font = Font(size=26)
    thin_border = Border(
        left=Side(border_style=BORDER_THICK, color='00000000'),
        right=Side(border_style=BORDER_THICK, color='00000000'),
        top=Side(border_style=BORDER_THICK, color='00000000'),
        bottom=Side(border_style=BORDER_THICK, color='00000000')
    )
    sheet.cell(1, 4).border = thin_border
    sheet.row_dimensions[1].height = float(43.00)

"""Add date format to insert date column"""
for sheet_name in wb.sheetnames:
    if sheet_name == "Events": continue
    sheet = wb[sheet_name]
    jcol = sheet['H']
    for row in jcol:
        row.number_format = "dd/mm/yy"

"""Add column headers"""
for sheet_name in wb.sheetnames:
    if sheet_name == "Events": continue
    sheet = wb[sheet_name]
    columnTitles = list(colNam.keys())
    sheet.insert_rows(2)
    for columnIndex, columnTitle in enumerate(columnTitles, start=1):
        sheet.cell(2, columnIndex).value = columnTitle

"""Add Dates on the top"""
for sheet_name in wb.sheetnames:
    if sheet_name == "Events": continue
    sheet = wb[sheet_name]
    date_cell = sheet.cell(1, 8).value
    if date_cell is not None:
        porcelain_date = np.busday_offset(date_cell.date(), -1, roll="backward", holidays=nz_holidays)
        sheet.cell(1, 1).value = pd.to_datetime(porcelain_date)
        sheet.cell(1, 1).font = Font(sz=26)
        sheet.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
        sheet.cell(1, 1).number_format = "dddd, d mmmm"
        sheet.merge_cells("A1:C1")

"""Adjust the column widths in all the sheets"""
for sheet_name in wb.sheetnames:
    if sheet_name == "Events": continue
    sheet = wb[sheet_name]
    for col, len in [('CaseID', 9), ('Dentist', 24), ('Patient', 22), ('Description', 26), ('Total', 5),
                     ('Last Department', 17), ('Ahead/Behind', 3),
                     ('INSERT DATE', 10)]:
        sheet.column_dimensions[colNam[col].letter].width = len

# """Create Dayily Total"""
# for sheet_name in wb.sheetnames:
#     if sheet_name == "Events": continue
#     sheet = wb[sheet_name]
#     x = colNam['Total'].letter
#     sheet.cell(1, 4).value = f"=sum({x}3:{x}{sheet.max_row})"
#     sheet.cell(1, 4).alignment = Alignment(horizontal="right")
#     sheet.cell(1, 4).font = Font(size=26)
#     sheet.merge_cells("D1:E1")

"""Conditional formatting"""
for sheet_name in wb.sheetnames:
    if sheet_name == "Events": continue
    sheet = wb[sheet_name]
    dxf = DifferentialStyle(fill=PatternFill(start_color='FFEAEC', end_color='FFEAEC'))
    # rule = Rule(type='expression', dxf=dxf, formula=[f"NOT(ISBLANK(${colNam['Problem'].letter}3))"])
    rule = Rule(type='expression', dxf=dxf, formula=[f"ISTEXT(${colNam['Total'].letter}3)"])
    x = colIdx[list(colIdx.keys())[-1]].letter
    sheet.conditional_formatting.add(f"A3:{x}{sheet.max_row + 1}", rule)

for sheet_name in wb.sheetnames:
    if sheet_name == "Events": continue
    sheet = wb[sheet_name]
    x = colNam["Ahead/Behind"].letter
    z = colNam["Total"].letter
    # last letter
    y = list(colLet.keys())[-1]

    # green for 2
    dxf = DifferentialStyle(fill=PatternFill(start_color='DBEFE0', end_color='DBEFE0'))
    rule = Rule(type='expression', dxf=dxf, formula=[f'=AND(${x}3>=2, ${z}3<>"Ignore List")'])
    f'=AND(${x}3=2, ${z}<>"Ignore; List")'
    sheet.conditional_formatting.add(f"A3:{y}{sheet.max_row + 1}", rule)

    if sheet_name in ["0", "1"]:
        # green for 1
        dxf = DifferentialStyle(fill=PatternFill(start_color='DBEFE0', end_color='DBEFE0'))
        rule = Rule(type='expression', dxf=dxf, formula=[f'=AND(${x}3=1, ${z}3<>"Ignore List")'])
        sheet.conditional_formatting.add(f"A3:{y}{sheet.max_row + 1}", rule)

    # yellow for -1
    dxf = DifferentialStyle(fill=PatternFill(start_color='FFF8D6', end_color='FFF8D6'))
    rule = Rule(type='expression', dxf=dxf, formula=[f'=AND(${x}3<=-1, ${z}3<>"Ignore List")'])
    sheet.conditional_formatting.add(f"A3:{y}{sheet.max_row + 1}", rule)

"""Other Totals at the bottom"""
max_rows = {}
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    a = colNam["Ahead/Behind"].letter
    t = colNam["Total"].letter
    ld = colNam["Last Department"].letter
    lr = sheet.max_row
    max_rows[sheet_name] = lr
    sheet.cell(lr + 1, 1).value = "Total"
    sheet.cell(lr + 1, 3).value = f"=sum({t}3:{t}{lr})"
    try:
        if sheet_name in ["5", "6", "7"]:
            m5 = int(sheet_name) - 5
            lr1 = max_rows[str(int(sheet_name) - 5)]
            sheet.cell(lr + 2, 1).value = "Waiting for parts"
            sheet.cell(lr + 2, 3).value = f"='{m5}'!C{lr1 + 6}"
    except KeyError:
        print("Probably a missing day")
    if sheet_name not in ["0", "1", "2"]:
        sheet.cell(lr + 3, 1).value = "Total ahead"
        sheet.cell(lr + 3, 3).value = f"=SUMIF({a}3:{a}{lr},\">=2\",{t}3:{t}{lr})"
        sheet.cell(lr + 4, 1).value = "Total minus ahead plus waiting for parts"
        sheet.cell(lr + 4, 3).value = f"=C{lr + 1}+C{lr + 2}-C{lr + 3}"
        continue
    sheet.cell(lr + 2, 1).value = "Total ahead"
    # for day 2 things that are 1 ahead don't matter (e.g. they are in sub structure when they are meant to be in models)
    if sheet_name == "2":
        sheet.cell(lr + 2, 3).value = f"=SUMIF({a}3:{a}{lr},\">=2\",{t}3:{t}{lr})"
    else:
        sheet.cell(lr + 2, 3).value = f"=SUMIF({a}3:{a}{lr},\">=1\",{t}3:{t}{lr})"
    sheet.cell(lr + 3, 1).value = "Total behind"
    sheet.cell(lr + 3, 3).value = f"=SUMIF({a}3:{a}{lr},\"<0\",{t}3:{t}{lr})"

    sheet.cell(lr + 4, 1).value = "Total behind -1"
    sheet.cell(lr + 4, 3).value = f"=SUMIFS({t}3:{t}{lr},{a}3:{a}{lr},\"=-1\",{ld}3:{ld}{lr},\"<>WAITING FOR PARTS\")"

    sheet.cell(lr + 5, 1).value = "Total behind -2"
    sheet.cell(lr + 5, 3).value = f"=SUMIFS({t}3:{t}{lr},{a}3:{a}{lr},\"=-2\",{ld}3:{ld}{lr},\"<>WAITING FOR PARTS\")"

    sheet.cell(lr + 6, 1).value = "Total Waiting for parts"
    sheet.cell(lr + 6, 3).value = f"=SUMIFS({t}3:{t}{lr},{ld}3:{ld}{lr},\"WAITING FOR PARTS\")"

    sheet.cell(lr + 7, 1).value = "Total minus ahead"
    sheet.cell(lr + 7, 3).value = f"=C{lr + 1}-C{lr + 2}"

    sheet.cell(lr + 8, 1).value = "Total minus ahead and behind"
    sheet.cell(lr + 8, 3).value = f"=C{lr + 1}-(C{lr + 2}+C{lr + 3})"

    try:
        if sheet_name in ["1", "2"]:
            rn1 = max_rows[str(int(sheet_name) - 1)]

            sheet.cell(lr + 8, 1).value = "Total minus A&B + Yest Behind"
            if sheet_name == "1":
                sheet.cell(lr + 8, 3).value = f"=C{lr + 7}+'{int(sheet_name) - 1}'!C{rn1 + 4}-C{lr+3}"
            elif sheet_name == "2":
                rn2 = max_rows[str(int(sheet_name) - 2)]
                sheet.cell(lr + 8,
                           3).value = f"=C{lr + 7}+'{int(sheet_name) - 1}'!C{rn1 + 4}+'{int(sheet_name) - 2}'!C{rn2 + 5}-C{lr+3}"
    except KeyError:
        print("Probably a missing day")

"""Make Summary Sheet"""
wb.create_sheet("Summary", 0)
summary_sheet = wb["Summary"]
summary_sheet.column_dimensions["A"].width = 2.68
summary_sheet.column_dimensions["B"].width = 10.0
summary_sheet.column_dimensions["D"].width = 2.68
summary_sheet.column_dimensions["F"].width = 46.0
summary_sheet.column_dimensions["H"].width = 2.68
summary_sheet.row_dimensions[1].height = 55

BonusPark = summary_sheet["B1"]
BonusPark.value = "Ballpark"
BonusPark.font = Font(name='Phosphate-Inline', size=27, )
BonusPark.alignment = Alignment(vertical="top")

TopDate = summary_sheet["F1"]
TopDate.value = "=TODAY()"
TopDate.font = Font(name='Phosphate-Inline', size=27)
TopDate.alignment = Alignment(vertical="top")
TopDate.number_format = "dd mmm"

row_counter = 0
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    lr = sheet.max_row
    if sheet_name == "0":
        continue
    try:
        sheet_name_num = int(sheet_name)
    except ValueError:
        # not one of the numbers
        continue
    thin_border_side = Side(border_style="thin", color="AAAAAA")
    top_bottom_border = Border(top=thin_border_side, bottom=thin_border_side, )

    summary_sheet.row_dimensions[2 + row_counter].height = 42.0
    date_cell = summary_sheet[f"B{2 + row_counter}"]
    date_cell.value = f"='{sheet_name}'!A1"
    date_cell.number_format = "ddd d"
    date_cell.font = Font(name='HelveticaNeue-Bold', size=13, )
    date_cell.alignment = Alignment(vertical="center")
    date_cell.border = Border(top=thin_border_side, bottom=thin_border_side, left=thin_border_side)

    total_used = summary_sheet[f"C{2 + row_counter}"]
    total_used.value = f"='{sheet_name}'!C{lr}"
    total_used.font = Font(name='HelveticaNeue', size=14, )
    total_used.alignment = Alignment(horizontal="right", vertical="center")
    total_used.border = top_bottom_border


    slash = summary_sheet[f"D{2 + row_counter}"]
    slash.value = "/"
    slash.font = Font(name='HelveticaNeue', size=14, )
    slash.alignment = Alignment(horizontal="center", vertical="center")
    slash.border = top_bottom_border

    total_capasity = summary_sheet[f"E{2 + row_counter}"]
    total_capasity.value = f"='{sheet_name}'!D1"
    total_capasity.font = Font(name='HelveticaNeue', size=14, )
    total_capasity.alignment = Alignment(vertical="center", horizontal="left")
    total_capasity.border = Border(top=thin_border_side, bottom=thin_border_side, right=thin_border_side)

    empty_bit = summary_sheet[f"F{2 + row_counter}"]
    empty_bit.border = Border(top=thin_border_side, bottom=thin_border_side, right=thin_border_side)

    rolling_total_used = summary_sheet[f"G{2+row_counter}"]
    rolling_total_used.value = f"=SUM(C2:C{2+row_counter})"
    rolling_total_used.font = Font(name='HelveticaNeue', size=14, italic=True, color="ACACAC" )
    rolling_total_used.alignment = Alignment(vertical="center", horizontal="right")
    rolling_total_used.number_format = "0"

    slash2 = summary_sheet[f"H{2+row_counter}"]
    slash2.value = "/"
    slash2.font = Font(name='HelveticaNeue', size=14, color="ACACAC" )
    slash2.alignment = Alignment(vertical="center", horizontal="center")

    rolling_total_capasity = summary_sheet[f"I{2+row_counter}"]
    rolling_total_capasity.value = f"=SUM(E2:E{2+row_counter})"
    rolling_total_capasity.font = Font(name='HelveticaNeue', size=14, italic=True, color="ACACAC" )
    rolling_total_capasity.alignment = Alignment(vertical="center", horizontal="left")

    # percentage = summary_sheet[f"J{2+row_counter}"]
    # percentage.value = f"=G{2+row_counter}/I{2+row_counter}"
    # percentage.font = Font(name='HelveticaNeue', size=14, )
    # percentage.number_format = numbers.FORMAT_PERCENTAGE
    # percentage.alignment = Alignment(vertical="center", horizontal="left")

    row_counter += 1

# Total at bottom
summary_sheet.row_dimensions[2 + row_counter].height = 35.0

thick_boarder = Side(border_style="thick", color="000001")
thing_border = Side(border_style="thin", color="000001")
thing_border_grey = Side(border_style="thin", color="AAAAAA")
top_border = Border(top=thick_boarder, bottom=thing_border)

cell = summary_sheet[f"B{2+row_counter}"]
cell.border = Border(top=thick_boarder, bottom=thing_border, left=thing_border_grey)

total_capasity_total = summary_sheet[f"C{2 + row_counter}"]
total_capasity_total.value = f"=SUM(C2:C{2 + row_counter -1})"
total_capasity_total.font = Font(name='HelveticaNeue', size=14)
total_capasity_total.alignment = Alignment(vertical="center", horizontal="right")
total_capasity_total.border = top_border

slash3 = summary_sheet[f"D{2 + row_counter}"]
slash3.value = "/"
slash3.font = Font(name='HelveticaNeue', size=14)
slash3.alignment = Alignment(vertical="center", horizontal="center")
slash3.border = top_border

rolling_total_capasity = summary_sheet[f"E{2 + row_counter}"]
rolling_total_capasity.value = f"=SUM(E2:E{2 + row_counter - 1})"
rolling_total_capasity.font = Font(name='HelveticaNeue', size=14)
rolling_total_capasity.alignment = Alignment(vertical="center", horizontal="left")
rolling_total_capasity.border = top_border

cell = summary_sheet[f"F{2+row_counter}"]
cell.border = Border(top=thick_boarder, bottom=thing_border, right=thing_border_grey)

dxf = DifferentialStyle(font=Font(color="C00001"))
rule = Rule(type='expression', dxf=dxf, formula=[f'=$C2>=$E2'])
summary_sheet.conditional_formatting.add(f"C2:E{2+row_counter-1}", rule)

wb.remove(wb['Events'])
wb.save(f"{load_wb[:-5]} Ballpark.xlsx")
print("Done")
